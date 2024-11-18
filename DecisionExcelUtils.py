import os
import re
import sys
import operator
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '.')))

import numpy as np
from copy import copy
from typing import Set, Dict 

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import range_boundaries, get_column_letter

from CreateDecisionTable import outputMsgBox
from logger import logger 
from ExcelConst import ExcelConst as const 

class DecisionExcelUtils:
    # ブックのオブジェクト
    wb = None
    # シートのオブジェクト
    ws = None
    # 既存フラグ
    exists_flag = False

    # コンストラクタ
    def __init__(self, template_path, sheet_name):
        # テンプレートファイルの存在チェック
        if not os.path.exists(template_path):
            raise FileNotFoundError(f'template_pathが存在しません')
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb[sheet_name]

        self.wb = wb
        self.ws = ws

    # コンストラクタ
    def __init__(self, template_path, sheet_name, exists_flag):
        # テンプレートファイルの存在チェック
        if not os.path.exists(template_path):
            raise FileNotFoundError(f'templateファイルが存在しません')
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb[sheet_name]

        self.wb = wb
        self.ws = ws
        self.exists_flag = exists_flag

    def get_min_to_max_coord(self, destinations):
        for title, coord in destinations:
            min_col, min_row, max_col, max_row = range_boundaries(coord)
        return min_col, min_row, max_col, max_row
    
    '''
        ケースを入力するエリアの編集を行う関数
        case_num
        ケース数
    '''

    def edit_case_input_area(self, case_num):
        temp_names = copy(self.wb.defined_names)
        for name, defined_name in temp_names.items():
            if const.AREA_NAME_CASEAREA not in defined_name.name:
                # 名前付き範囲が「ケースエリア」以外の場合はスキップ
                continue
            # 最小・最大の行番号・列番号を取得
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(list(defined_name.destinations))

            # 行数を取得
            col_num = max_col - min_col + 1

            if case_num <= col_num:
                raise Exception(f'リクエスト用のパラメータファイルに記載しているNoが既に存在しています')
            
            if col_num < case_num:
                # ケースを記載する列の列数がケース数よりも小さい場合
                add_num = case_num - col_num
                # 処理するエリアを取得
                target_area = list(self.ws.iter_rows(min_col=max_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=False))
                # 行を追加
                for i in range(0, add_num):
                    self.add_cols(target_area, max_col + 1, True)
    
    '''
        ASTAIDやURLを記載するエリアを追加する関数
        api_num
        apiの呼び出し回数
    '''
    
    def add_title_area(self, api_num):
        # 2行目からコピー
        title_area = list(self.ws.iter_rows(min_row=const.TITLE_MIN_ROW, max_row=const.TITLE_MAX_ROW, values_only=False))

        # 行を追加
        # api_numはAPIの呼び出し回数
        # テンプレートファイルには1回目のIDやURL等のエリアがあるため、api_numを-1回する
        for i in range(0, api_num -1):
            # self.ws:シートのオブジェクト
            # source:コピー元の行単位のデータ
            # 5 + add_row_num:5は5行目に追加
            # 追加した行数
            add_row_num = i * len(title_area)
            # タイトルを記載するエリアを追加
            self.add_rows(title_area, 5 + add_row_num, False)
            # 追加した行数
            # APIのリクエストとレスポンスを記載している行をコピー
            # 名前付き範囲を取得
            named_range = self.wb.defined_names[f'{const.AREA_NAME_API}{i+1}']
            # 範囲の境界を取得
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(named_range.destinations)
            # APIのリクエストやレスポンスを記載するエリアを取得
            api_area = list(self.ws.iter_rows(min_row=min_row, max_row=max_row, values_only=False))
            # 追加する行数
            add_row_num = (i * len(api_area)) + ((i+1) * len(title_area))
            # リクエストとレスポンスを記載するエリアを追加
            self.add_rows(api_area, 39 + add_row_num, True)

    '''
        名前付き範囲の行数えお追加・削除する処理
        param_list
        名前付き範囲に記載するパラメータの一覧
        target_area_name
        対象の名前付き範囲
        padding_num
        追加する位置（0の場合名前付き範囲の最初の行の位置から行を追加・削除する）
    '''

    def increase_decrease_named_area_lines(self, param_list, target_area_name, padding_num):
        temp = copy(self.wb.defined_names)
        for name, defined_name in temp.items():
            if target_area_name not in defined_name.name:
                # リクエストパラメータを記載する行以外はスキップ
                continue

            # 名前付き範囲「リクエストパラメータエリア」の末尾の数字を取得
            # 末尾の数字は何回目のAPIのパラメータかを示している
            result = re.match(r'(\D+)(\d+)', name)
            if result:
                pattern1, pattern2 = result.groups()
                param_idx = int(pattern2)
                if param_idx > 0:
                    param_idx = param_idx - 1

            else:
                logger.error(f'テンプレートファイルの名前付き範囲が更新されています、確認してください')
                outputMsgBox('エラーがあります、ログを確認してください','エラー')
                return
            
            # 範囲の境界を取得
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(list(defined_name.destinations))
            
            param_size = 0
            if const.AREA_NAME_REQUEST in target_area_name:
                # リクエストパラメータの場合の処理
                for key, value_list in param_list[param_idx].items():
                    for value in value_list:
                        param_size += 1
            elif const.AREA_NAME_RESPONSE in target_area_name:
                # レスポンスパラメータの場合の処理
                for api_res_object in param_list[param_idx]:
                    for key, value in api_res_object.items():
                        if key != "":
                            param_size += 1
            
            # 行を追加
            row_num = max_row - min_row + 1

            if row_num < param_size:
                add_num = param_size - row_num
                # 処理するエリアを取得
                target_area = list(self.ws.iter_rows(min_row=min_row+padding_num, max_row=min_row+padding_num, values_only=False))
                
                # 行を追加
                for i in range(0, add_num):
                    self.add_rows(target_area, min_row + padding_num, False)
            elif row_num > param_size:
                delete_num = row_num - param_size
                # 処理するエリアを取得
                target_area = list(self.ws.iter_rows(min_row=min_row+padding_num, max_row=min_row+padding_num, values_only=False))
                
                # 行を削除
                for i in range(0, delete_num):
                    self.delete_rows(target_area, min_row)
    
    '''
    リクエストパラメータの設定をする関数
    param_list
    パラメータリスト
    target_area_name
    名前付き範囲
    '''
    
    def edit_request_param(self, param_list, request_data, target_area_name):
        for name, defined_name in self.wb.defined_names.items():
            if target_area_name not in defined_name.name:
                # target_area_nameの行はスキップ
                continue

            # 名前付き範囲「リクエストパラメータエリア」の末尾の数字を取得    
            # 末尾の数字は何回目のAPIのパラメータかを示している
            result = re.match(r'(\D+)(\d+)', name)
            if result:
                pattern1, pattern2 = result.groups()
                param_idx = int(pattern2)
                if param_idx > 0:
                    param_idx = param_idx - 1
            else:
                logger.error(f'テンプレートファイルの名前付き範囲が更新されています、確認してください')
                outputMsgBox('エラーがあります、ログを確認してください','エラー')
                return
            
            # 最小行と最大行を取得
            named_range = self.wb.defined_names[defined_name.name]
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(named_range.destinations)

            # パラメータの総数を取得
            param_size = 0
            for key, value_list in param_list[param_idx].items():
                for value in value_list:
                    param_size += 1
            
            # キーと値のリストを作成
            key_list = []
            value_list = []
            for key, values in param_list[param_idx].items():
                for value in values:
                    key_list.append(key)
                    value_list.append(value)
            
            # マージ後のリクエストパラメータ数文、ループ
            for i in range(0, len(key_list)):
                # キー名を設定
                key_cell = self.ws.cell(row=i + min_row, column=const.INPUT_KEY_COL)
                key_cell.value = key_list[i]
                # 値を設定
                value_cell = self.ws.cell(row=i + min_row, column=const.INPUT_VALUE_COL)
                value_cell.value = value_list[i]

                for case_num, case_object in enumerate(request_data, start=0):
                    if request_data[case_num][param_idx] is None:
                        continue
                    for key, values in request_data[case_num][param_idx].items():
                        for value in values:
                            if key == key_list[i] and value == value_list[i]:
                                # ●付け
                                decision_cell = self.ws.cell(row=i + min_row, column=const.INPUT_AREA_COL + case_num)
                                decision_cell.value ='●'
    
    '''
    レスポンスパラメータの設定する関数
    param_list
    パラメータリスト
    target_area_name
    名前付き範囲
    '''
    def edit_response_param(self, param_list, target_area_name):
        temp = copy(self.wb.defined_names)
        for name, defined_name in temp.items():
            if target_area_name not in defined_name.name:
                # target_area_nameの行以外はスキップ
                continue

            # 名前付き範囲「リクエストパラメータエリア」の末尾の数字を取得
            # 末尾の数字は何回目のAPIのパラメータかを示している
            result = re.match(r'(\D+)(\d+)', name)
            if result:
                pattern1, pattern2 = result.groups()
                param_idx = int(pattern2)
                if param_idx > 0:
                    param_idx = param_idx - 1
            else:
                logger.error(f'テンプレートファイルの名前付き範囲が更新されています、確認してください')
                outputMsgBox('エラーがあります、ログを確認してください','エラー')
                return
            
            # 最小行と最大行を取得
            named_range = self.wb.defined_names[defined_name.name]
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(named_range.destinations)

            # ケース数分、ループ処理
            param_cnt = 0
            for case_num, api_object in enumerate(param_list[param_idx], start=0):
                if len(api_object) == 0:
                    # レスポンスがないため、次のケースのレスポンスを処理
                    continue
                # レスポンスパラメータ数文、ループ処理
                for key, value in api_object.items():

                    if key == '$.status':
                        # KEY名が「＄.status」の場合、テンプレートファイルに存在しているので、
                        # テンプレートファイルの該当箇所に●付け
                        if value[0] == self.ws.cell(row=min_row -2, column=min_col +3).value:
                            # SUCCESSの場合
                            self.ws.cell(row=min_row -2, column=const.INPUT_AREA_COL + case_num).value = '●'
                            # レスポンスパラメータを記載する行を1行削除
                            self.delete_line(min_row + param_cnt)

                            continue
                        elif value[0] == self.ws.cell(row=min_row -1, column=min_col +3).value:
                            # VALIDATION_FAILUREの場合
                            self.ws.cell(row=min_row -1, column=const.INPUT_AREA_COL + case_num).value = '●'
                            # レスポンスパラメータを記載する行を1行削除
                            self.delete_line(min_row + param_cnt)

                            continue
                    
                    # key名を設定
                    self.ws.cell(row=min_row + param_cnt, column=const.INPUT_KEY_COL).value = key
                    # valueを設定
                    self.ws.cell(row=min_row + param_cnt, column=const.INPUT_VALUE_COL).value = value[0]
                    # ●付け
                    self.ws.cell(row=min_row + param_cnt, column=const.INPUT_AREA_COL + case_num).value = '●'
                    param_cnt += 1
    
    def delete_line(self, row):
        # 処理するエリアを取得
        target_area = list(self.ws.iter_rows(min_row=row+1, max_row=row+1, values_only=False))
        # 行削除
        self.delete_rows(target_area, row)

    '''
    行追加により入力規則の範囲がずれる場合、
    本関数を使用して入力規則の範囲を追加した行数分ずらす関数
    source
    コピー元オブジェクト
    edit
        col:列に対して操作を行う場合
        row:行に対して操作を行う場合
    mode
        delete:削除する場合
        add:追加する場合
    '''
    def translate_data_validations(self, source, edit, mode):
        # コピー元の最小行
        src_min_row = source[0][0].row
        # コピー元の最大行
        src_max_row = source[len(source)-1][0].row
        # コピー元の最小列
        src_min_col = source[0][0].column
        # コピー元の最大列
        src_max_col = source[0][len(source[0])-1].column

        # 要素数が変わる可能性があり、ループ処理の途中で変わってしまうとエラーになる
        # ループ自体はコピーしたオブジェクトを使用
        temp_data_validation = copy(self.ws.data_validations.dataValidation)
        for dv in temp_data_validation:
            new_sqref = []
            temp_ranges = copy(dv.sqref.ranges)
            for ref in temp_ranges:
                min_col, min_row, max_col, max_row = range_boundaries(ref.coord)
                if edit == 'row':
                    if src_min_row <= min_row and src_max_row <= min_row:
                        # コピー元の範囲が名前付き範囲より上の行にある場合、名前付き範囲の最小行・最大行を更新
                        if src_min_row <= ref.min_row and src_max_row <= ref.min_row:
                            if mode == 'add':
                                ref.min_row += len(source)
                                ref.max_row += len(source)
                            elif mode == 'delete':
                                ref.min_row -= len(source)
                                ref.max_row -= len(source)
                    elif src_min_row >= min_row and min_row < src_max_row <= max_row:
                        if src_min_row >= ref.min_row and ref.min_row < src_max_row <= ref.max_row:
                            if mode == 'add':
                                ref.max_row += len(source)
                            elif mode == 'delete':
                                ref.max_row -= len(source)
                elif edit == 'col':
                    if src_min_col <= min_col and src_max_col <= min_col:
                        # コピー元の範囲が名前付き範囲より上の行にある場合、名前付き範囲の最小列・最大列を更新
                        if src_min_col <= ref.min_col and src_max_col <= ref.min_col:
                            col_num = src_max_col - src_min_col + 1
                            if mode == 'add':
                                ref.min_col += col_num
                                ref.max_col += col_num
                            elif mode == 'delete':
                                ref.min_col -= col_num
                                ref.max_col -= col_num
                    elif src_min_col >= min_col and min_col < src_max_col <= max_col:
                        if src_min_col >= ref.min_col and ref.min_col < src_max_col <= ref.max_col:
                            col_num = src_max_col - src_min_col + 1
                            if mode == 'add':
                                ref.max_col += col_num
                            elif mode == 'delete':
                                ref.max_col -= col_num
    
    '''
    行追加を行う際にコピー元の行に入力規則が存在している場合、
    本関数を使用して入力規則の範囲をコピーして新しく追加した行に入力規則を適用させる関数
    source
    コピー元オブジェクト
    edit
        col:列に対して操作を行う場合
        row:行に対して操作を行う場合
    '''
    def add_data_validations(self, source, edit):
        # データの入力規則をコピー
        # 要素数が変わる可能性があり、ループ処理の途中で変わってしまうとエラーになる
        # ループ自体はコピーしたオブジェクトを使用
        temp_data_validation = copy(self.ws.data_validations.dataValidation)

        min_col = 0
        min_row = 0
        max_col = 0
        max_row = 0

        add_col_num = source[0][len(source[0]) -1].column - source[0][0].column + 1

        # 入力規則をループ処理
        for dv in temp_data_validation:
            temp_ranges = copy(dv.ranges)
            # 入力規則の範囲をループ処理
            for ref in temp_ranges.ranges:
                # コピーした列数分、ループ処理
                add_flag = False
                for row in range(source[0][0].row, source[len(source) - 1][0].row):
                    for col in range(source[0][0].column, source[0][len(source[0]) - 1].column + 1):
                        cell = self.ws.cell(row=row, column=col)
                        if cell.coordinate in ref:

                            if not add_flag and edit == 'row':
                                min_col = col
                                min_row = row + len(source)
                                max_col = col
                                max_row = row + len(source)
                                add_flag = True
                                continue
                            elif not add_flag and edit == 'col':
                                min_col = col
                                min_row = row
                                max_col = col + add_col_num
                                max_row = row 
                                add_flag = True
                                continue
                                
                        if edit == 'row':
                            # 列の最小・最大の更新
                            if min_col >= col:
                                min_col = col
                            if max_col <= col:
                                max_col = col
                            # 行の最小・最大の更新
                            if min_row >= row + len(source):
                                min_row = row + len(source)
                            if max_row <= row + len(source):
                                max_row = row + len(source)

                        elif edit == 'col':
                            # 行の最小・最大の更新
                            if min_col >= col + add_col_num:
                                min_col = col + add_col_num
                            if max_row <= col + add_col_num:
                                max_row = col + add_col_num
                            # 行の最小・最大の更新
                            if min_row >= row:
                                min_row = row
                            if max_row <= row:
                                max_row = row
                if add_flag:
                    dv.add(f'${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}')

    
    '''
    名前付き範囲の更新処理を行う関数
    source
    コピー元オブジェクト
    edit_line_flag
        True:末尾に対して追加を行う場合
        False:途中に追加する場合
    edit
        col:列に対して操作を行う場合
        row:行に対して操作を行う場合
    mode
        delete:削除する場合
        add:追加する場合
    '''

    def edit_named_range(self, source, end_line_flag, edit, mode):
        # 名前付き範囲の処理
        temp_defined_names = copy(self.wb.defined_names)
        for name, defined_name in temp_defined_names.items():
            # 名前付き範囲がテーブル形式の場合はスキップ
            match = re.search(r'.*\[.*\].*', defined_name.attr_text)
            if match:
                continue

            destinations = list(defined_name.destinations)
            for title, coord in destinations:
                # 名前付き範囲の境界を取得
                min_col, min_row, max_col, max_row = range_boundaries(coord)
                # 追加する行が何行目かを取得
                src_min_row = source[0][0].row
                src_max_row = source[len(source) -1][0].row
                src_min_col = source[0][0].column
                src_max_col = source[0][len(source[0]) -1].column

                if edit == 'row' and not end_line_flag:
                    new_coord = self.calc_rows_range(min_row, max_row, min_col, max_col, src_min_row, src_max_row, mode)
                    self.wb.defined_names[name].attr_text = f'{new_coord}'
                elif edit == 'col' and not end_line_flag:
                    # 名前付き範囲を名前と末尾の数字に分割
                    new_coord = self.calc_cols_range(min_row, max_row, min_col, max_col, src_min_row, src_max_row, mode)
                    self.wb.defined_names[name].attr_text = f'{new_coord}'
                elif end_line_flag:
                    result = re.match(r'(\D+)(\d+)', name)
                    if edit == 'row' and result:
                        pattern1, pattern2 = result.groups()
                        pattern2 = int(pattern2) + 1
                        new_name = pattern1 + str(pattern2)
                    else:
                        new_name = name
                    
                    if edit == 'row':
                        if not(new_name == 'テストケースNoエリア' or new_name == 'ケースエリア' or new_name == '●付けエリア'):
                            if mode == 'add':
                                new_coord = f'{self.ws.title}!${get_column_letter(min_col)}${min_row + len(source)}:${get_column_letter(max_col)}${max_row + len(source)}'
                            else:
                                new_coord = f'{self.ws.title}!${get_column_letter(min_col)}${min_row - len(source)}:${get_column_letter(max_col)}${max_row - len(source)}'
                            self.wb.defined_names.add(DefinedName(new_name, attr_text=new_coord))
                        elif new_name == 'ケースエリア' or new_name == '●付けエリア':
                            new_coord = f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row + len(source)}'
                            self.wb.defined_names[new_name].attr_text = new_coord
                    elif edit == 'col':
                        if new_name in self.wb.defined_names:
                            col_num = src_max_col - src_min_col + 1
                            self.wb.defined_names[new_name].attr_txet = f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col + col_num)}${max_row}'

    '''
    名前付き範囲の行の範囲を計算する関数
    min_row
    参照中の名前付き範囲の最小行
    max_row
    参照中の名前付き範囲最大行
    src_min_row
    追加する行の最小行
    src_max_row
    追加する行の最大行
    mode
        'add':行追加
        'minus':行削除

    '''
    def calc_rows_range(self, min_row, max_row, min_col, max_col, src_min_row, src_max_row, mode):
        row_num = src_max_row - src_min_row + 1
        if min_row <= src_min_row and src_max_row <= max_row:
            # 追加する行が名前付き範囲の範囲内の場合、
            # 名前付き範囲の最大行を追加する行数だけシフトさせる
            if mode =='add':
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row + row_num}'
            elif mode == 'delete' and min_row <= (max_row - row_num):
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row - row_num}'
            elif mode == 'delete' and min_row > (max_row - row_num):
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}'
        elif not(max_row < src_min_row and max_row < src_max_row):
            # 名前付き範囲と被っている場合、もしくは
            # 名前付き範囲の最小行：最小行 + 行数
            # 名前付き範囲の最大行；最大行 + 行数
            if mode == 'add':
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row + row_num}:${get_column_letter(max_col)}${max_row + row_num}'
            elif mode == 'delete' and min_row <= (max_row - row_num):
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row - row_num}:${get_column_letter(max_col)}${max_row - row_num}'
            elif mode == 'delete' and min_row > (max_row - row_num):
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}'
    
        return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}'

    '''
    名前付き範囲の列の範囲を計算する関数
    min_row
    参照中の名前付き範囲の最小行
    max_row
    参照中の名前付き範囲最大行
    src_min_row
    追加する行の最小行
    src_max_row
    追加する行の最大行
    mode
        'add':行追加
        'minus':行削除

    '''
    
    def calc_cols_range(self, min_row, max_row, min_col, max_col, src_min_col, src_max_col, mode):
        col_num = src_max_col - src_min_col
        if min_col <= src_min_col and src_max_col <= max_col:
            # 追加する列が名前付き範囲の範囲内の場合、
            # 名前付き範囲の最大列を追加する列数だけシフトさせる
            if mode == 'add':
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col + col_num)}${max_row}'
            else:
                return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col - col_num)}${max_row}'
        else:
            # 名前付き範囲と被っている場合、
            # 名前付き範囲の最小列：最小列 + 列数
            # 名前付き範囲の最大列：最大列 + 列数
            if mode == 'add':
                return f'{self.ws.title}!${get_column_letter(min_col + col_num)}${min_row}:${get_column_letter(max_col + col_num)}${max_row}'
            else:
                return f'{self.ws.title}!${get_column_letter(min_col - col_num)}${min_row}:${get_column_letter(max_col - col_num)}${max_row}'
            
        return f'{self.ws.title}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}'


    '''
    行を追加するための関数
    source
    コピー元のオブジェクト
    start_row
    行追加する位置
    end_line_flag
        True:末尾に対して追加を行う場合
        False:途中に追加する場合
    '''
    
    def add_rows(self, source, start_row, end_line_flag):
        # 行を挿入
        self.ws.insert_rows(start_row, amount=len(source))

        if not end_line_flag:
            # 行追加によりデータ入力規則の範囲がずれるため、範囲の更新を行う
            self.translate_data_validations(source, 'row', 'add')
            
            # 元の行の高さを保持
            row_heights = {}
            for row in range(start_row, self.ws.max_row + len(source)):
                row_heights[row] = self.ws.row_dimensions[row].height

                # 挿入してずれた行の高さを再設定
                for row, height in row_heights.items():
                    self.ws.row_dimensions[row + len(source)].height = height
        else:

            # 末尾に行を追加するため、コピー元に入力規則がある場合に入力規則をコピー
            self.add_data_validations(source, 'row')

        # 挿入した行に対してコピー元に入力規則がある場合に入力規則をコピー
        for row_idx, row in enumerate(source, start=start_row):
            # コピー元の高さを取得
            source_height = self.ws.row_dimensions[row[0].row].height
            # コピー先の高さを設定
            self.ws.row_dimensions[row_idx].height = source_height
            for col_idx, cell in enumerate(row, start=1):
                target = self.ws.cell(row=row_idx, column=col_idx)
                # コピー元の値を設定
                if not self.exists_flag:
                    target.value = cell.value
                # コピー元のスタイルを設定
                target._style = copy(cell._style)

        # 名前付き範囲の更新
        self.edit_named_range(source, end_line_flag, 'row', 'add')
    
    '''
    行を削除するための関数
    source
    コピー元オブジェクト
    start_row
    行追加する位置
    end_line_flag
        True:末尾に対して追加を行う場合
        False:途中に追加する場合
    '''
    def delete_rows(self, source, start_row):
        # 行を追加
        self.ws.delete_rows(start_row, len(source))

        # 行追加によりデータ入力規則の範囲がずれるため、範囲の更新を行う
        self.translate_data_validations(source, 'row', 'delete')

        # 名前付き範囲の更新
        self.edit_named_range(source, False, 'row','delete')
    
    '''
    列を追加するための関数
    source
    コピー元オブジェクト
    start_col
    列追加する位置
    end_line_flag
        True:末尾に対して追加を行う場合
        False:途中に追加する場合
    '''
    def add_cols(self, source, start_col, end_line_flag):

        min_col = source[0][0].column
        max_col = source[0][len(source[0]) - 1].column

        add_col_num = max_col - min_col + 1

        # 列を挿入
        self.ws.insert_cols(start_col, amount=add_col_num)
        self.translate_data_validations(source, 'col', 'add')

        # if not end_line_flag:
        #   # 行追加によりデータ入力規則の範囲がずれるため、範囲の更新を行う
        #   self.translate_data_validations(source, 'col', 'add')
        # else:
        #   # 
        #   self.add_data_validations(source, 'col')

        for col_idx in range(start_col, start_col + add_col_num):
            # コピー先の高さを設定
            self.ws.column_dimensions[get_column_letter(col_idx)].width = 8.5
        
        # 挿入した行に対してコピー元のデータとスタイルを設定
        for copy_cells in source:
            for cell in copy_cells:
                target = self.ws.cell(row=cell.row, column=cell.column + add_col_num)
                # コピー元の値を設定
                if not self.exists_flag:
                    target.value = cell.value
                # コピー元のスタイルを設定
                target._style = copy(cell._style)
        
        # 名前付き範囲の更新
        self.edit_named_range(source, end_line_flag, 'col', 'add')
    
    '''
    ケースNoとスクリプトNoを入力する関数
    case_num
    ケース数
    api_num
    APIの呼び出し回数
    '''
    def input_case_no(self, case_num, api_num):
        # テストケースNoを記載している行を取得
        title_end_row = (api_num * 3) + 1 # ASTAIDやURL等を記載しているエリアの最終行を設定
        case_no_row = title_end_row + 5 # テストケースNoの行
        script_no_row = title_end_row + 6 # スクリプトNoの行 

        # ループのカウントを1からカウントさせるため、case_num(ケース数)も+1させる
        for case_no in range(1, case_num + 1):
            # テストケースNoの行とスクリプトNoの行を処理
            for row in range(case_no_row, script_no_row + 1):
                cell = self.ws.cell(row=row, column=const.INPUT_AREA_COL + case_no -1)
                cell.value = case_no
    
    '''
    正常・異常を入力するエリアの表示ルールを設定する関数
    case_num
    ケース数
    api_num
    APIの呼び出し回数
    '''
    def create_case_condition_rules(self, case_num, api_num):
        rule1 = CellIsRule(operator='equal', formula=['"正常"'],fill=PatternFill(start_color='92d050', end_color='92d050', fill_type='solid'))
        rule2 = CellIsRule(operator='equal', formula=['"正常"'],fill=PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'))
        # ルールを適用
        row = 9 + (api_num -1) * 3
        row -= 1
        self.ws.conditional_formatting.add(f'{get_column_letter(const.INPUT_AREA_COL)}{row}:{get_column_letter(const.INPUT_AREA_COL + case_num +1)}{row}', rule1)
        self.ws.conditional_formatting.add(f'{get_column_letter(const.INPUT_AREA_COL)}{row}:{get_column_letter(const.INPUT_AREA_COL + case_num +1)}{row}', rule2)
    
    '''
    ケース数のチェック
    '''
    def check_case_count(self, case_num):
        for name, defined_name in self.wb.defined_names.items():
            if const.AREA_NAME_CASEAREA not in defined_name.name:
                # target_area_nameの行以外はスキップ
                continue

            # 最小行と最大行を取得
            named_range = self.wb.defined_names[defined_name.name]
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(named_range.destinations)

            count = max_col - min_col +1
            if case_num <= count:
                raise Exception(f'既存のディシジョンテーブルに既に同じケースNoのケースが存在している可能性があります')
            return count
    
    '''
    リクエストパラメータの取得
    '''
    def get_param_list(self, case_num, api_num, param_list, target_name):
        for name, defined_name in self.wb.defined_names.items():
            if target_name not in defined_name.name:
                # target_area_nameの行以外はスキップ
                continue

            # 名前付き範囲を名前と末尾の数字に分割
            result = re.match(r'(\D+)(\d+)', name)
            if result:
                pattern1, pattern2 = result.groups()
                api_idx = int(pattern2) -1 # 何回目に呼び出されるAPIを判別するための番号
            
            # 最小行と最大行を取得
            named_range = self.wb.defined_names[name]
            min_col, min_row, max_col, max_row = self.get_min_to_max_coord(named_range.destinations)
            # ケース数分、ループ処理
            for i in range(0, case_num):
                data : Dict[str, Set[int]] = {}
                # api_idx何回目のリクエストパラメータ数分、ループ処理
                for row in range(min_row, max_row +1):
                    # keyの取得
                    key = self.ws.cell(row=row, column=const.INPUT_KEY_COL).value
                    # valueの取得
                    value = self.ws.cell(row=row, column=const.INPUT_VALUE_COL).value
                    
                    if '●' == self.ws.cell(row=row, column=const.INPUT_AREA_COL + i).value:
                        if target_name == const.AREA_NAME_REQUEST:
                            if key not in data:
                                data[key] = set()
                            data[key].add(value)
                        else:
                            if key not in data:
                                data[key]=''
                            data[key] = value
                            
                # リクエストパラメータリストに値を追加
                param_list[i][api_idx] = data
        
        return param_list
    
    '''
    ●付けしたエリアを初期化
    '''
    def clear_decision_area(self, target_list):
        for target in target_list:
            for name, deifined_name in self.wb.defined_names.items():
                if target not in deifined_name.name:
                    # target_area_nameの行以外はスキップ
                    continue
                # 最小行と最大行を取得
                named_range = self.wb.defined_names[deifined_name.name]
                cells = named_range.destinations

                for title, coord in cells:
                    for row in self.ws[coord]:
                        for cell in row:
                            cell.value = None
    
    '''
    エクセルを保存する関数
    path
    フルパス
    '''
    def save_book(self, path):
        self.wb.save(path)



